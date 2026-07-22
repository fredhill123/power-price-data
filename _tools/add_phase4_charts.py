"""
add_phase4_charts.py — Phase-4 chart surgery on the query-wired deliverable
workbook (HourlyPowerData.xlsx), preserving all Power Query parts byte-for-byte.

Adds 6 charts:
  chart10  Iberia (Spain) intraday price shape        clone chart2, Fig2_Intraday DE->ES
  chart11  Germany intraday duck curve (absolute)     clone chart2, sheet->Fig2_Intraday_avg (DE)
  chart12  Portugal capture price vs baseload         clone chart6, Fig5_Capture DE->PT
  chart13  Iberia (Spain) cumulative near-neg hours   clone chart4, Fig3_CumNeg DE->ES
  chart14  Solar peak-hour share (DE/ES/PT, qavg)     NEW empty tab G1_SolarPeak  (Fred wires query)
  chart15  Germany intraday duck by month (2025)      NEW empty tab G2_MonthDuck  (Fred wires query)

chart10-13 anchor on the existing 'Charts' tab (drawing10) below chart9.
chart14/15 anchor on their own new empty tabs (empty-target pattern) appended at
the END of the sheet list, so every existing localSheetId (0-based doc-order index
of the ExternalData_1 scoped names) is untouched.

Never opens the workbook with openpyxl for WRITING (would strip PQ). openpyxl is
used data_only for READING loaded cell values only.
"""
from __future__ import annotations
import os, re, zipfile, warnings
from lxml import etree
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
warnings.filterwarnings("ignore")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Clean base = Fred's query-wired file BEFORE any Phase-4 charts (idempotent re-runs).
SRC  = os.path.join(ROOT, "archive", "phase4_2026-07-17", "HourlyPowerData_pre-phase4.xlsx")
OUT  = os.path.join(ROOT, "outputs", "HourlyPowerData.xlsx")
CH   = os.path.join(ROOT, "outputs", "csv", "charts")

C  = "http://schemas.openxmlformats.org/drawingml/2006/chart"
A  = "http://schemas.openxmlformats.org/drawingml/2006/main"
RNS= "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
def c(t): return f"{{{C}}}{t}"
def a(t): return f"{{{A}}}{t}"

# Redburn palette
NAVY="2E3E80"; TEAL="5FA1AD"; SAGE="ACBFB7"; FOREST="3D664A"; GOLD="CC9F53"; WINE="8A1E41"

# ---- read source zip preserving member order -------------------------------
zin = zipfile.ZipFile(SRC)
parts = {i.filename: zin.read(i.filename) for i in zin.infolist()}
order = [i.filename for i in zin.infolist()]
zin.close()
wbv = openpyxl.load_workbook(SRC, data_only=True)   # READ-ONLY, for cached values

from completeness import cutoffs
LCY = cutoffs()["last_complete_year"]               # single-year charts track the latest complete year
FIG67_BASE_YEAR = 2024                              # the year the base workbook's Fig6/Fig7 charts ship at

def ser_root(name):  # parse a chart part
    return etree.fromstring(parts[name])

def dump(root):
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

def num_v(v):
    return repr(float(v)) if isinstance(v, (int, float)) else str(v)

# ---- reference helpers -----------------------------------------------------
_REF = re.compile(r'\$([A-Z]+)\$(\d+)')
def shift_ref(ftext, offset, new_sheet=None):
    """Shift every column ref (except col A = category) by `offset`; optionally
    rename the sheet. Handles 'Sheet!$B$2:$B$25' and single '$B$1'."""
    sheet, rng = ftext.split("!", 1)
    if new_sheet: sheet = new_sheet
    def repl(m):
        col, row = m.group(1), m.group(2)
        if col == "A":
            return f"$A${row}"
        return f"${get_column_letter(column_index_from_string(col) + offset)}${row}"
    return f"{sheet}!{_REF.sub(repl, rng)}"

def build_num_cache(values, fmt="General"):
    nc = etree.Element(c("numCache"))
    etree.SubElement(nc, c("formatCode")).text = fmt
    etree.SubElement(nc, c("ptCount")).set("val", str(len(values)))
    for i, v in enumerate(values):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        pt = etree.SubElement(nc, c("pt")); pt.set("idx", str(i))
        etree.SubElement(pt, c("v")).text = num_v(v)
    return nc

def build_str_cache(values):
    sc = etree.Element(c("strCache"))
    etree.SubElement(sc, c("ptCount")).set("val", str(len(values)))
    for i, v in enumerate(values):
        if v is None:
            continue
        pt = etree.SubElement(sc, c("pt")); pt.set("idx", str(i))
        etree.SubElement(pt, c("v")).text = str(v)
    return sc

def cells(sheet, col_letter, r0, r1):
    ws = wbv[sheet]
    return [ws[f"{col_letter}{r}"].value for r in range(r0, r1 + 1)]

# ---------------------------------------------------------------------------
# A. Country-variant clones (chart10-13)
# ---------------------------------------------------------------------------
def clone_variant(template, offset, cache_sheet, new_sheet=None, yaxis=None):
    """Clone a Redburn chart, shift columns by `offset`, rebuild each series'
    value numCache from `cache_sheet`, optionally rename the sheet + y-axis title."""
    root = ser_root(template)
    for ser in root.iter(c("ser")):
        # --- value ref + cache ---
        vf = ser.find(f"{c('val')}/{c('numRef')}/{c('f')}")
        vf.text = shift_ref(vf.text, offset, new_sheet)
        m = re.search(r'!\$([A-Z]+)\$(\d+):\$[A-Z]+\$(\d+)', vf.text)
        col, r0, r1 = m.group(1), int(m.group(2)), int(m.group(3))
        numref = vf.getparent()
        numref.replace(numref.find(c("numCache")), build_num_cache(cells(cache_sheet, col, r0, r1)))
        # --- category ref (col A -> just rename sheet if asked; cache unchanged) ---
        cat = ser.find(c("cat"))
        if cat is not None:
            cf = cat.find(f"{c('numRef')}/{c('f')}")
            if cf is None:
                cf = cat.find(f"{c('strRef')}/{c('f')}")
            cf.text = shift_ref(cf.text, offset, new_sheet)
    if yaxis:
        t = root.find(f".//{c('valAx')}/{c('title')}//{a('t')}")
        if t is not None:
            t.text = yaxis
    return dump(root)

parts["xl/charts/chart10.xml"] = clone_variant("xl/charts/chart2.xml", 17, "Fig2_Intraday")            # DE->ES
parts["xl/charts/chart11.xml"] = clone_variant("xl/charts/chart2.xml", 0,  "Fig2_Intraday_avg",
                                                new_sheet="Fig2_Intraday_avg", yaxis="€/MWh")      # DE absolute
parts["xl/charts/chart12.xml"] = clone_variant("xl/charts/chart6.xml", 34, "Fig5_Capture")             # DE->PT
parts["xl/charts/chart13.xml"] = clone_variant("xl/charts/chart4.xml", 17, "Fig3_CumNeg")              # DE->ES

# ---------------------------------------------------------------------------
# B. New-tab charts (chart14 G1, chart15 G2-monthly) — clone chart2, multiply series
# ---------------------------------------------------------------------------
def ramp(colors, n):
    """Interpolate a list of hex anchors to n colours."""
    def h2r(h): return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    def r2h(r): return "".join(f"{max(0,min(255,int(round(x)))):02X}" for x in r)
    anch = [h2r(x) for x in colors]
    if n == 1: return [colors[0]]
    out = []
    for k in range(n):
        pos = k / (n - 1) * (len(anch) - 1)
        lo = int(pos); hi = min(lo + 1, len(anch) - 1); f = pos - lo
        out.append(r2h(tuple(anch[lo][j] + (anch[hi][j] - anch[lo][j]) * f for j in range(3))))
    return out

def build_multiseries(sheet, cat_col, cat_vals, cat_is_str, r0, series, yaxis):
    """series = list of (name, col_letter, values, hexcolor). Returns chart XML."""
    root = ser_root("xl/charts/chart2.xml")
    line = root.find(f".//{c('lineChart')}")
    protos = line.findall(c("ser"))
    proto = protos[0]
    for s in protos:            # strip template series
        line.remove(s)
    # rebuild category element (shared shape) once as template
    rN = r0 + len(cat_vals) - 1
    dLbls = line.find(c("dLbls"))   # keep dLbls/smooth/axId AFTER series
    for i, (name, col, vals, color) in enumerate(series):
        ser = etree.fromstring(etree.tostring(proto))   # deep copy
        ser.find(c("idx")).set("val", str(i))
        ser.find(c("order")).set("val", str(i))
        ser.find(f"{c('tx')}/{c('v')}").text = name
        ser.find(f"{c('spPr')}/{a('ln')}/{a('solidFill')}/{a('srgbClr')}").set("val", color)
        # category
        cat = ser.find(c("cat"))
        for ch in list(cat): cat.remove(ch)
        if cat_is_str:
            ref = etree.SubElement(cat, c("strRef"))
            etree.SubElement(ref, c("f")).text = f"{sheet}!${cat_col}${r0}:${cat_col}${rN}"
            ref.append(build_str_cache(cat_vals))
        else:
            ref = etree.SubElement(cat, c("numRef"))
            etree.SubElement(ref, c("f")).text = f"{sheet}!${cat_col}${r0}:${cat_col}${rN}"
            ref.append(build_num_cache(cat_vals))
        # value
        val = ser.find(c("val"))
        for ch in list(val): val.remove(ch)
        vref = etree.SubElement(val, c("numRef"))
        etree.SubElement(vref, c("f")).text = f"{sheet}!${col}${r0}:${col}${rN}"
        vref.append(build_num_cache(vals))
        line.insert(list(line).index(dLbls), ser)
    # y-axis title
    t = root.find(f".//{c('valAx')}/{c('title')}//{a('t')}")
    if t is not None: t.text = yaxis
    # keep category (date/hour) labels at the bottom even when values go negative (C erosion)
    tlp = root.find(f".//{c('catAx')}/{c('tickLblPos')}")
    if tlp is not None: tlp.set("val", "low")
    return dump(root)

# --- G1: solar peak-hour share, quarterly-avg lines DE/ES/PT ---
g1 = pd.read_csv(os.path.join(CH, "g1_solar_peakhour.csv"))
g1_dates = g1["date"].tolist()                       # 4383 rows, dates present through 2030
G1_SERIES = [("Germany", "C", g1["DE_qavg"].tolist(), NAVY),
             ("Spain",   "E", g1["ES_qavg"].tolist(), WINE),
             ("Portugal","G", g1["PT_qavg"].tolist(), GOLD)]
parts["xl/charts/chart14.xml"] = build_multiseries(
    "G1_SolarPeak", "A", g1_dates, True, 2, G1_SERIES, "solar share of peak hour (%)")

# --- G2: Germany intraday price by month, 2025 (duck curve) ---
g2 = pd.read_csv(os.path.join(CH, "g2_price_by_month.csv"))
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
g2cols = list(g2.columns)
g2_hours = g2["hour_utc"].tolist()                   # 24 rows
mcolors = ramp([NAVY, TEAL, GOLD, WINE], 12)
G2_SERIES = []
for mi, mon in enumerate(MONTHS, start=1):
    col_name = f"DE_{LCY}_M{mi:02d}"
    col_letter = get_column_letter(g2cols.index(col_name) + 1)
    G2_SERIES.append((mon, col_letter, g2[col_name].tolist(), mcolors[mi-1]))
parts["xl/charts/chart15.xml"] = build_multiseries(
    "G2_MonthDuck", "A", g2_hours, False, 2, G2_SERIES, "€/MWh")

# ---------------------------------------------------------------------------
# B2. Charts 16-19 — monthly "market-state" charts on their own new empty tabs
#     (empty-target: PQ loads figA/B/C/D CSVs into $A$2.. on the Windows box)
# ---------------------------------------------------------------------------
_CO_COLS = [("Germany", "B", "DE", NAVY), ("Spain", "C", "ES", WINE),
            ("Portugal", "D", "PT", GOLD), ("France", "E", "FR", TEAL), ("Italy", "F", "IT", FOREST)]

# chart16 — A: monthly baseload price by country
figA = pd.read_csv(os.path.join(CH, "figA_monthly_price.csv"))
A_dates = figA["date"].tolist()
A_SERIES = [(nm, col, figA[cc].tolist(), clr) for nm, col, cc, clr in _CO_COLS]
parts["xl/charts/chart16.xml"] = build_multiseries("A_MonthPrice", "A", A_dates, True, 2, A_SERIES, "€/MWh")

# chart17 — B: wind+solar penetration (12-mo avg) by country
figB = pd.read_csv(os.path.join(CH, "figB_penetration.csv"))
B_dates = figB["date"].tolist()
B_SERIES = [(nm, col, figB[cc].tolist(), clr) for nm, col, cc, clr in _CO_COLS]
parts["xl/charts/chart17.xml"] = build_multiseries(
    "B_Penetration", "A", B_dates, True, 2, B_SERIES, "wind + solar, % of generation")

# chart18 — C: solar/wind capture price vs baseload (Germany)
figC = pd.read_csv(os.path.join(CH, "figC_capture_erosion.csv"))
C_dates = figC["date"].tolist()
C_SERIES = [("Solar", "B", figC["DE_Solar"].tolist(), GOLD),
            ("Onshore wind", "C", figC["DE_Wind"].tolist(), TEAL)]
parts["xl/charts/chart18.xml"] = build_multiseries(
    "C_CaptureErosion", "A", C_dates, True, 2, C_SERIES, "capture vs baseload, %")

# chart19 — D: net-load "duck" deepening (demand - wind - solar), Germany, by year
figD = pd.read_csv(os.path.join(CH, "figD_netload_duck.csv"))
D_hours = figD["hour_utc"].tolist()
D_ycols = [col for col in figD.columns if col.startswith("DE_") and figD[col].notna().any()]
D_years = [int(col.split("_")[1]) for col in D_ycols]
D_colors = ramp(["C9D2CD", TEAL, NAVY], len(D_years))            # matches charts.year_colors ramp
D_SERIES = []
for i, (col, y) in enumerate(zip(D_ycols, D_years)):
    letter = get_column_letter(list(figD.columns).index(col) + 1)
    lab = f"{y} YTD" if y > LCY else str(y)
    D_SERIES.append((lab, letter, figD[col].tolist(), D_colors[i]))
parts["xl/charts/chart19.xml"] = build_multiseries(
    "D_NetloadDuck", "A", D_hours, False, 2, D_SERIES, "net load, GW (demand − wind − solar)")
print("built chart16-19 (monthly market-state)")

# --- repoint base Fig6 (chart7) + Fig7 (chart8) single-year charts to the latest complete year ---
def repoint_year(chart_name, sheet, col_shift):
    root = ser_root(chart_name)
    for tag in ("val", "xVal", "yVal"):
        for ref in root.iter(c(tag)):
            f = ref.find(f"{c('numRef')}/{c('f')}")
            if f is None:
                continue
            f.text = shift_ref(f.text, col_shift)      # skips col A (category) automatically
            m = re.search(r'!\$([A-Z]+)\$(\d+):\$[A-Z]+\$(\d+)', f.text)
            col, r0, r1 = m.group(1), int(m.group(2)), int(m.group(3))
            numref = f.getparent()
            nc = numref.find(c("numCache"))
            if nc is not None:
                numref.replace(nc, build_num_cache(cells(sheet, col, r0, r1)))
    return dump(root)

_sy = LCY - FIG67_BASE_YEAR
if _sy:
    parts["xl/charts/chart7.xml"] = repoint_year("xl/charts/chart7.xml", "Fig6_MinMax", _sy * 1)
    parts["xl/charts/chart8.xml"] = repoint_year("xl/charts/chart8.xml", "Fig7_GenMix", _sy * 20)

    # The DATA moved to LCY above; the human-readable year has to move with it, or the
    # workbook shows LCY figures under a base-year caption. Two places carry it:
    #   1. the Charts-tab captions for Fig 6 / Fig 7, inherited from the base workbook
    #   2. chart7's single series name ("DE 2024"), a literal with no cell behind it
    # Both were still on FIG67_BASE_YEAR until 2026-07-22.
    _s14 = parts["xl/worksheets/sheet14.xml"].decode()
    _before = _s14
    for _cap in ("Daily minimum vs maximum price (Germany, ",
                 "Intraday generation mix and price (Portugal, "):
        _s14 = _s14.replace(f"{_cap}{FIG67_BASE_YEAR})", f"{_cap}{LCY})")
    if _s14 == _before:
        raise SystemExit(f"!! Fig6/Fig7 captions not found at {FIG67_BASE_YEAR} on the Charts "
                         f"tab — the base workbook's wording changed. Fix this mapping rather "
                         f"than shipping charts whose caption year contradicts their data.")
    parts["xl/worksheets/sheet14.xml"] = _s14.encode()

    _c7 = parts["xl/charts/chart7.xml"].decode()
    _c7, _n7 = re.subn(rf">DE {FIG67_BASE_YEAR}<", f">DE {LCY}<", _c7)
    if not _n7:
        raise SystemExit(f"!! chart7's series name is no longer 'DE {FIG67_BASE_YEAR}' — "
                         f"check what it is before assuming the year rolled.")
    parts["xl/charts/chart7.xml"] = _c7.encode()
    print(f"repointed Fig6/Fig7 {FIG67_BASE_YEAR}->{LCY} (data, captions and series name)")

# --- label the partial (current) year "YTD" in profile charts, matching the static path ---
def relabel_ytd(key):
    root = ser_root(key)
    changed = False
    for tx in root.iter(c("tx")):
        v = tx.find(c("v"))
        if v is not None and v.text and v.text.strip().isdigit() and int(v.text.strip()) > LCY:
            v.text = f"{v.text.strip()} YTD"; changed = True
    if changed:
        parts[key] = dump(root)
for _cn in ("chart2", "chart4", "chart5", "chart10", "chart11", "chart13"):   # the year-profile charts
    relabel_ytd(f"xl/charts/{_cn}.xml")

print("built chart10-15")

# ---------------------------------------------------------------------------
# C. Anchor chart10-13 on the 'Charts' tab (drawing10) + captions on sheet14
# ---------------------------------------------------------------------------
XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
def anchor_xml(col, row, rid, shape_id, name):
    return (f'<xdr:oneCellAnchor xmlns:xdr="{XDR}" xmlns:a="{A}">'
            f'<xdr:from><xdr:col>{col}</xdr:col><xdr:colOff>0</xdr:colOff>'
            f'<xdr:row>{row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
            f'<xdr:ext cx="6120000" cy="3240000"/>'
            f'<xdr:graphicFrame macro=""><xdr:nvGraphicFramePr>'
            f'<xdr:cNvPr id="{shape_id}" name="{name}"/><xdr:cNvGraphicFramePr/></xdr:nvGraphicFramePr>'
            f'<xdr:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></xdr:xfrm>'
            f'<a:graphic><a:graphicData uri="{C}">'
            f'<c:chart xmlns:c="{C}" xmlns:r="{RNS}" r:id="{rid}"/>'
            f'</a:graphicData></a:graphic></xdr:graphicFrame><xdr:clientData/></xdr:oneCellAnchor>')

# grid continues after chart9 (col0/row81): 10=col11/row81, 11=col0/row101, 12=col11/row101, 13=col0/row121
NEW_ON_CHARTS = [  # (chartnum, col, row, caption)
    (10, 11, 81,  "Spain — intraday price shape, indexed to daily mean"),
    (11, 0, 101,  "Germany — intraday price shape (€/MWh), the deepening 'duck' belly"),
    (12, 11, 101, "Portugal — capture price vs baseload by technology"),
    (13, 0, 121,  "Spain — cumulative near-negative-price hours through the year"),
]
d10 = parts["xl/drawings/drawing10.xml"].decode()
d10r = parts["xl/drawings/_rels/drawing10.xml.rels"].decode()
add_anchor = ""
add_rel = ""
shape_id = 11   # last existing = 10 (Chart 9)
for k, (cn, col, row, cap) in enumerate(NEW_ON_CHARTS):
    rid = f"rId{10 + k}"     # existing rId1..9 used
    add_anchor += anchor_xml(col, row, rid, shape_id + k, f"Chart {cn}")
    add_rel += (f'<Relationship Id="{rid}" '
                f'Type="{RNS}/chart" Target="../charts/chart{cn}.xml"/>')
parts["xl/drawings/drawing10.xml"] = d10.replace("</xdr:wsDr>", add_anchor + "</xdr:wsDr>").encode()
parts["xl/drawings/_rels/drawing10.xml.rels"] = d10r.replace("</Relationships>", add_rel + "</Relationships>").encode()

# captions on sheet14 (Charts) — inlineStr, navy bold Arial, at the from-col/row of each chart
def caption_cell(col_letter, row, text):
    return (f'<c r="{col_letter}{row}" t="inlineStr"><is><r>'
            f'<rPr><b/><sz val="12"/><color rgb="FF2E3E80"/><rFont val="Arial"/></rPr>'
            f'<t xml:space="preserve">{text}</t></r></is></c>')
COLMAP = {0: "A", 11: "L"}
s14 = parts["xl/worksheets/sheet14.xml"].decode()
# insert caption cells as their own rows just above each chart's from-row (row index is 0-based in drawing -> +1 for 1-based cell row; captions sit at row = drawing_row (i.e. the row above chart top))
cap_rows = ""
for cn, col, row, cap in NEW_ON_CHARTS:
    cell_row = row            # drawing 'row' is 0-based; the 1-based row number == row (chart top is at row+1)
    cap_rows += f'<row r="{cell_row}">{caption_cell(COLMAP[col], cell_row, cap)}</row>'
s14 = s14.replace("</sheetData>", cap_rows + "</sheetData>")
parts["xl/worksheets/sheet14.xml"] = s14.encode()
print("anchored chart10-13 on Charts tab")

# ---------------------------------------------------------------------------
# D. New empty sheets G1_SolarPeak / G2_MonthDuck (chart14/15) appended at END
# ---------------------------------------------------------------------------
def empty_sheet_xml():
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'xmlns:r="{RNS}"><dimension ref="A1"/><sheetViews><sheetView workbookViewId="0"/></sheetViews>'
            '<sheetFormatPr defaultRowHeight="15"/><sheetData/><drawing r:id="rId1"/></worksheet>').encode()

def sheet_drawing_rels(drawing_no):
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'<Relationship Id="rId1" Type="{RNS}/drawing" Target="../drawings/drawing{drawing_no}.xml"/>'
            '</Relationships>').encode()

def new_drawing_xml(chart_rid):
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<xdr:wsDr xmlns:xdr="{XDR}" xmlns:a="{A}"><xdr:oneCellAnchor>'
            '<xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff>'
            '<xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
            '<xdr:ext cx="8100000" cy="4400000"/>'
            '<xdr:graphicFrame macro=""><xdr:nvGraphicFramePr>'
            '<xdr:cNvPr id="2" name="Chart 1"/><xdr:cNvGraphicFramePr/></xdr:nvGraphicFramePr>'
            '<xdr:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></xdr:xfrm>'
            f'<a:graphic><a:graphicData uri="{C}">'
            f'<c:chart xmlns:c="{C}" xmlns:r="{RNS}" r:id="{chart_rid}"/>'
            '</a:graphicData></a:graphic></xdr:graphicFrame><xdr:clientData/></xdr:oneCellAnchor></xdr:wsDr>').encode()

def new_drawing_rels(chart_no):
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'<Relationship Id="rId1" Type="{RNS}/chart" Target="../charts/chart{chart_no}.xml"/>'
            '</Relationships>').encode()

NEW_SHEETS = [  # (name, worksheet_no, drawing_no, chart_no, sheetId, rId) — all appended at END
    ("G1_SolarPeak", 15, 11, 14, 15, "rId21"),
    ("G2_MonthDuck", 16, 12, 15, 16, "rId22"),
    ("A_MonthPrice", 17, 13, 16, 17, "rId23"),
    ("B_Penetration", 18, 14, 17, 18, "rId24"),
    ("C_CaptureErosion", 19, 15, 18, 19, "rId25"),
    ("D_NetloadDuck", 20, 16, 19, 20, "rId26"),
]
for name, wsno, dno, cno, sid, rid in NEW_SHEETS:
    parts[f"xl/worksheets/sheet{wsno}.xml"] = empty_sheet_xml()
    parts[f"xl/worksheets/_rels/sheet{wsno}.xml.rels"] = sheet_drawing_rels(dno)
    parts[f"xl/drawings/drawing{dno}.xml"] = new_drawing_xml("rId1")
    parts[f"xl/drawings/_rels/drawing{dno}.xml.rels"] = new_drawing_rels(cno)
    order += [f"xl/worksheets/sheet{wsno}.xml", f"xl/worksheets/_rels/sheet{wsno}.xml.rels",
              f"xl/drawings/drawing{dno}.xml", f"xl/drawings/_rels/drawing{dno}.xml.rels"]

# workbook.xml: append the two <sheet> entries (END -> preserves existing localSheetId indices)
wbxml = parts["xl/workbook.xml"].decode()
add_sheets = "".join(f'<sheet name="{n}" sheetId="{sid}" r:id="{rid}"/>'
                     for n, _, _, _, sid, rid in NEW_SHEETS)
wbxml = wbxml.replace("</sheets>", add_sheets + "</sheets>")
parts["xl/workbook.xml"] = wbxml.encode()

# workbook.xml.rels: worksheet rels for the two new sheets
wbrels = parts["xl/_rels/workbook.xml.rels"].decode()
add_wbrel = "".join(f'<Relationship Id="{rid}" Type="{RNS}/worksheet" Target="worksheets/sheet{wsno}.xml"/>'
                    for _, wsno, _, _, _, rid in NEW_SHEETS)
wbrels = wbrels.replace("</Relationships>", add_wbrel + "</Relationships>")
parts["xl/_rels/workbook.xml.rels"] = wbrels.encode()
print("added G1_SolarPeak + G2_MonthDuck sheets")

# ---------------------------------------------------------------------------
# E. Content types: charts 10-15, drawings 11/12, worksheets 15/16
# ---------------------------------------------------------------------------
ct = parts["[Content_Types].xml"].decode()
adds = ""
for n in range(10, 20):
    adds += (f'<Override PartName="/xl/charts/chart{n}.xml" '
             'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>')
for dno in (11, 12, 13, 14, 15, 16):
    adds += (f'<Override PartName="/xl/drawings/drawing{dno}.xml" '
             'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
for wsno in (15, 16, 17, 18, 19, 20):
    adds += (f'<Override PartName="/xl/worksheets/sheet{wsno}.xml" '
             'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
ct = ct.replace("</Types>", adds + "</Types>")
parts["[Content_Types].xml"] = ct.encode()

# register new chart parts in the zip order
for n in range(10, 20):
    order.append(f"xl/charts/chart{n}.xml")

# ---- write out -------------------------------------------------------------
os.makedirs(os.path.dirname(OUT), exist_ok=True)
# de-dup order while preserving first occurrence
seen = set(); final_order = []
for nm in order:
    if nm not in seen:
        seen.add(nm); final_order.append(nm)
zout = zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED)
for nm in final_order:
    zout.writestr(nm, parts[nm])
zout.close()
print(f"wrote {OUT}  ({len(final_order)} parts)")
