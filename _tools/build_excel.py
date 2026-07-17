"""
build_excel.py — write the fixed-cell-reference summary workbook.

Design contract (critical: PowerPoint charts link to these cells):
  * Every datapoint sits at a DETERMINISTIC cell = f(country, year, hour, tech,
    doy, pct) — independent of the data values and of which years have data yet.
  * Time-series tabs pre-allocate year columns out to config.DISPLAY_END_YEAR
    (blank until data arrives) so linked charts auto-extend into future years
    with no cell movement. Single-panel tabs (Fig6/Fig7) order blocks YEAR-then-
    country so a new year appends at the bottom without shifting existing cells.
  * One worksheet per Redburn figure, so each chart links to one clean range.

Reads : data/processed/summaries/*.parquet
Writes: output/PowerPriceData.xlsx
"""
from __future__ import annotations
import os, warnings; warnings.filterwarnings("ignore")
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config as cfg

SUM = os.path.join(cfg.PROC_DIR, "summaries")
OUT = os.path.join(cfg.OUTPUT_DIR, "PowerPriceData.xlsx")

# Time-series tabs pre-allocate the full display grid (out to DISPLAY_END_YEAR)
# so linked charts auto-extend into future years with no cell movement.
YEARS = cfg.DISPLAY_YEARS
# Single-panel tabs (Fig6/Fig7) only lay out years that have data; a new year
# appends cleanly at the bottom on the next rebuild.
DATA_YEARS = cfg.YEARS
CUR = cfg.CURRENT_YEAR
COUNTRIES = cfg.COUNTRY_ORDER
TECHS = cfg.TECH_ORDER

# styles
NAVY = "1F3864"; TEAL = "2E7D8A"; GREY = "F2F2F2"
H1 = Font(bold=True, color="FFFFFF", size=12)
H2 = Font(bold=True, color="FFFFFF", size=9)
BOLD = Font(bold=True)
FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_TEAL = PatternFill("solid", fgColor=TEAL)
FILL_GREY = PatternFill("solid", fgColor=GREY)
CENTER = Alignment(horizontal="center")
YTD = PatternFill("solid", fgColor="FFF2CC")  # highlight partial year

def _load(name):
    return pd.read_parquet(os.path.join(SUM, f"{name}.parquet"))

def _title(ws, text, sub=""):
    ws["A1"] = text; ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    if sub:
        ws["A2"] = sub; ws["A2"].font = Font(italic=True, size=9, color="808080")

def _country_year_header(ws, r0, c0, metric_note=""):
    """Two-row header: country (merged over its years) then year. Returns dict
    (country,year)->column index. Marks 2026 as YTD."""
    colmap = {}
    c = c0
    for country in COUNTRIES:
        ws.cell(r0, c, cfg.COUNTRIES[country]["name"]).font = H2
        ws.cell(r0, c).fill = FILL_NAVY; ws.cell(r0, c).alignment = CENTER
        ws.merge_cells(start_row=r0, start_column=c, end_row=r0, end_column=c + len(YEARS) - 1)
        for y in YEARS:
            cell = ws.cell(r0 + 1, c, y); cell.font = H2; cell.fill = FILL_TEAL
            cell.alignment = CENTER
            if y == CUR:
                cell.fill = YTD; cell.font = BOLD
            colmap[(country, y)] = c
            c += 1
    return colmap

# ---------------------------------------------------------------------------
def sheet_readme(wb, updated):
    ws = wb.active; ws.title = "README"
    _title(ws, "ENTSO-E Power Price Data — summary workbook")
    lines = [
        ("Last updated (UTC)", updated),
        ("Coverage", f"Germany (DE-LU), Spain, Portugal, France, Italy — {cfg.START_YEAR} to {CUR}"),
        (f"{CUR}", "CURRENT YEAR = PARTIAL / year-to-date (highlighted amber where shown)"),
        ("Timezone", "ALL analytics bucketed by UTC hour (DST-safe, per spec)"),
        ("Italy price", "Load-weighted PUN proxy across bidding zones (labelled)"),
        ("Units", "Price EUR/MWh · generation & flows MW (hourly mean) · capacity MW"),
        ("Cell stability", "Every datapoint has a FIXED cell reference across regenerations. "
                            f"Time-series tabs pre-allocate year columns to {cfg.DISPLAY_END_YEAR} "
                            "(blank until data arrives) so linked PowerPoint charts auto-extend into "
                            "future years with NO cell movement. Fig6/Fig7 append new years at the bottom."),
        ("Future-proof", f"Pipeline auto-fetches new years (currently to {CUR}); rerun refresh.sh each period."),
        ("", ""),
        ("Tab", "Figure / content"),
        ("Fig1_PriceSD", "Std-dev of hourly price distribution, per year (Redburn Fig 1)"),
        ("Fig2_IntradayPrice", "Indexed & avg price by UTC hour, per year (Fig 2 + requested)"),
        ("Fig3_NegHours", "Negative & near-negative hour counts per year (Fig 3 totals)"),
        ("Fig3_CumNegHours", "Cumulative near-neg hours by day-of-year (Fig 3 curves)"),
        ("Fig4_DurationCurve", "Price duration curve by percentile, per year (Fig 4)"),
        ("Fig5_Capture", "Capture price vs base % and absolute, per tech per year (Fig 5)"),
        ("Fig6_DailyMinMax", "Daily min/max/mean price by day-of-year (Fig 6 scatter)"),
        ("Fig7_GenMix", "Avg intraday generation mix + net flow + price by UTC hour (Fig 7)"),
        ("Fig9_Capacity", "Annual installed capacity by technology (Fig 9)"),
        ("CaptureMonthly", "Capture price per tech per month (requested analytic)"),
    ]
    r = 4
    for a, b in lines:
        ws.cell(r, 1, a).font = BOLD if a and not b.startswith(("Std","Indexed","Cumul","Price","Capture","Daily","Avg","Annual","Neg")) else Font()
        ws.cell(r, 1, a)
        ws.cell(r, 2, b)
        r += 1
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 95

def sheet_price_sd(wb):
    ws = wb.create_sheet("Fig1_PriceSD")
    _title(ws, "Fig 1 — Std-dev of hourly power-price distribution", "EUR/MWh · by year · UTC")
    r0 = 4
    ws.cell(r0, 1, "Year").font = BOLD
    for j, c in enumerate(COUNTRIES):
        ws.cell(r0, 2 + j, cfg.COUNTRIES[c]["name"]).font = H2
        ws.cell(r0, 2 + j).fill = FILL_NAVY; ws.cell(r0, 2 + j).alignment = CENTER
    df = _load("price_sd").set_index(["country", "year"])["sd"]
    for i, y in enumerate(YEARS):
        rr = r0 + 1 + i
        cell = ws.cell(rr, 1, y); cell.font = BOLD
        if y == CUR: cell.fill = YTD
        for j, c in enumerate(COUNTRIES):
            v = df.get((c, y))
            ws.cell(rr, 2 + j, round(float(v), 2) if pd.notna(v) else None)
    ws.column_dimensions["A"].width = 8
    for j in range(len(COUNTRIES)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 11

def sheet_intraday_price(wb):
    ws = wb.create_sheet("Fig2_IntradayPrice")
    _title(ws, "Fig 2 — Indexed & average power price by UTC hour",
           "Indexed (1 = annual base) and avg EUR/MWh · per country per year")
    df = _load("intraday_price")
    r = 4
    for country in COUNTRIES:
        ws.cell(r, 1, cfg.COUNTRIES[country]["name"] + " — INDEXED (1 = annual mean)").font = Font(bold=True, color=NAVY)
        hdr = r + 1
        ws.cell(hdr, 1, "Hour UTC").font = BOLD
        for j, y in enumerate(YEARS):
            cell = ws.cell(hdr, 2 + j, y); cell.font = H2; cell.fill = FILL_TEAL; cell.alignment = CENTER
            if y == CUR: cell.fill = YTD
        sub = df[df.country == country]
        for h in range(24):
            rr = hdr + 1 + h
            ws.cell(rr, 1, f"{h:02d}h")
            for j, y in enumerate(YEARS):
                v = sub[(sub.year == y) & (sub.hour_utc == h)]["indexed"]
                ws.cell(rr, 2 + j, round(float(v.iloc[0]), 4) if len(v) and pd.notna(v.iloc[0]) else None)
        # avg block to the right
        c_off = 2 + len(YEARS) + 1
        ws.cell(hdr, c_off - 1, "avg EUR/MWh →").font = Font(italic=True, size=8)
        for j, y in enumerate(YEARS):
            cell = ws.cell(hdr, c_off + j, y); cell.font = H2; cell.fill = FILL_TEAL; cell.alignment = CENTER
            if y == CUR: cell.fill = YTD
        for h in range(24):
            rr = hdr + 1 + h
            for j, y in enumerate(YEARS):
                v = sub[(sub.year == y) & (sub.hour_utc == h)]["avg_price"]
                ws.cell(rr, c_off + j, round(float(v.iloc[0]), 2) if len(v) and pd.notna(v.iloc[0]) else None)
        r = hdr + 1 + 24 + 2  # next country block (fixed 29-row stride)
    ws.column_dimensions["A"].width = 14

def sheet_neg_hours(wb):
    ws = wb.create_sheet("Fig3_NegHours")
    _title(ws, "Fig 3 — Negative & near-negative price hours per year",
           "count of hours · negative = price<0 · near-negative = price<€1/MWh · UTC")
    df = _load("neg_hours").set_index(["country", "year"])
    for bi, (metric, lbl) in enumerate([("neg_hours", "NEGATIVE (price < 0)"),
                                        ("near_neg_hours", "NEAR-NEGATIVE (price < €1/MWh)")]):
        r0 = 4 + bi * (len(YEARS) + 4)
        ws.cell(r0, 1, lbl).font = Font(bold=True, color=NAVY)
        hdr = r0 + 1
        ws.cell(hdr, 1, "Year").font = BOLD
        for j, c in enumerate(COUNTRIES):
            ws.cell(hdr, 2 + j, cfg.COUNTRIES[c]["name"]).font = H2
            ws.cell(hdr, 2 + j).fill = FILL_NAVY; ws.cell(hdr, 2 + j).alignment = CENTER
        for i, y in enumerate(YEARS):
            rr = hdr + 1 + i
            cell = ws.cell(rr, 1, y); cell.font = BOLD
            if y == CUR: cell.fill = YTD
            for j, c in enumerate(COUNTRIES):
                v = df["".join([metric])].get((c, y)) if (c, y) in df.index else None
                ws.cell(rr, 2 + j, int(v) if pd.notna(v) else None)
    ws.column_dimensions["A"].width = 10
    for j in range(len(COUNTRIES)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 11

def sheet_cum_neg(wb):
    ws = wb.create_sheet("Fig3_CumNegHours")
    _title(ws, "Fig 3 — Cumulative near-negative hours by day-of-year",
           "cumulative count of hours with price < €1/MWh · UTC")
    df = _load("cum_neghours")
    r0 = 4
    colmap = _country_year_header(ws, r0, 2)
    ws.cell(r0 + 1, 1, "Day of year").font = BOLD
    for d in range(1, 367):
        rr = r0 + 2 + (d - 1)
        ws.cell(rr, 1, d)
        for (c, y), col in colmap.items():
            v = df[(df.country == c) & (df.year == y) & (df.doy == d)]["cum_near_neg"]
            ws.cell(rr, col, int(v.iloc[0]) if len(v) and pd.notna(v.iloc[0]) else None)
    ws.column_dimensions["A"].width = 11

def sheet_duration(wb):
    ws = wb.create_sheet("Fig4_DurationCurve")
    _title(ws, "Fig 4 — Annual price duration curves",
           "price (EUR/MWh) at each percentile of hours (0% = highest-priced hour) · UTC")
    df = _load("duration_curve")
    r0 = 4
    colmap = _country_year_header(ws, r0, 2)
    ws.cell(r0 + 1, 1, "% of hours").font = BOLD
    pcts = sorted(df["pct_of_hours"].unique())
    for i, pct in enumerate(pcts):
        rr = r0 + 2 + i
        ws.cell(rr, 1, pct)
        for (c, y), col in colmap.items():
            v = df[(df.country == c) & (df.year == y) & (df.pct_of_hours == pct)]["price"]
            ws.cell(rr, col, round(float(v.iloc[0]), 2) if len(v) and pd.notna(v.iloc[0]) else None)
    ws.column_dimensions["A"].width = 10

def sheet_capture(wb):
    ws = wb.create_sheet("Fig5_Capture")
    _title(ws, "Fig 5 — Capture price by technology vs base price",
           "% above/below base price (top block) and absolute EUR/MWh (lower block) · annual · UTC")
    df = _load("capture_annual")
    for bi, (metric, lbl, rnd) in enumerate([
            ("capture_vs_base_pct", "% ABOVE / BELOW BASE PRICE", 1),
            ("capture_price", "ABSOLUTE CAPTURE PRICE (EUR/MWh)", 2)]):
        r0 = 4 + bi * (len(TECHS) + 5)
        ws.cell(r0, 1, lbl).font = Font(bold=True, color=NAVY)
        hdr = r0 + 1
        colmap = _country_year_header(ws, hdr, 2)
        ws.cell(hdr + 1, 1, "Technology").font = BOLD
        for i, t in enumerate(TECHS):
            rr = hdr + 2 + i
            ws.cell(rr, 1, t)
            for (c, y), col in colmap.items():
                v = df[(df.country == c) & (df.year == y) & (df.tech == t)][metric]
                ws.cell(rr, col, round(float(v.iloc[0]), rnd) if len(v) and pd.notna(v.iloc[0]) else None)
    ws.column_dimensions["A"].width = 26

def sheet_capacity(wb):
    ws = wb.create_sheet("Fig9_Capacity")
    _title(ws, "Fig 9 — Installed generation capacity by technology", "MW · annual · source ENTSO-E")
    df = _load("capacity")
    r0 = 4
    colmap = _country_year_header(ws, r0, 2)
    ws.cell(r0 + 1, 1, "Technology").font = BOLD
    for i, t in enumerate(TECHS):
        rr = r0 + 2 + i
        ws.cell(rr, 1, t)
        for (c, y), col in colmap.items():
            v = df[(df.country == c) & (df.year == y) & (df.tech == t)]["capacity_mw"]
            ws.cell(rr, col, round(float(v.iloc[0]), 1) if len(v) and pd.notna(v.iloc[0]) else None)
    ws.column_dimensions["A"].width = 26

def sheet_daily_minmax(wb):
    ws = wb.create_sheet("Fig6_DailyMinMax")
    _title(ws, "Fig 6 — Daily min / max / mean price by day-of-year",
           "EUR/MWh · single-panel blocks ordered YEAR then country (a new year appends "
           "at the bottom without shifting existing blocks) · fixed 370-row stride · UTC")
    df = _load("daily_minmax").copy()
    df["doy"] = pd.to_datetime(df["date"]).dt.dayofyear
    STRIDE = 370
    r = 4
    for y in DATA_YEARS:                      # YEAR outer -> future years append at bottom
        for country in COUNTRIES:
            ws.cell(r, 1, f"{cfg.COUNTRIES[country]['name']} {y}").font = Font(bold=True, color=NAVY)
            if y == CUR: ws.cell(r, 1).fill = YTD
            hdr = r + 1
            for k, lab in enumerate(["Day of year", "Min", "Max", "Mean", "Spread"]):
                ws.cell(hdr, 1 + k, lab).font = H2; ws.cell(hdr, 1 + k).fill = FILL_NAVY
            sub = df[(df.country == country) & (df.year == y)].set_index("doy")
            for d in range(1, 367):
                rr = hdr + d
                ws.cell(rr, 1, d)
                if d in sub.index:
                    row = sub.loc[d]
                    ws.cell(rr, 2, row["min_price"]); ws.cell(rr, 3, row["max_price"])
                    ws.cell(rr, 4, row["mean_price"]); ws.cell(rr, 5, row["spread"])
            r += STRIDE
    ws.column_dimensions["A"].width = 16

def sheet_genmix(wb):
    ws = wb.create_sheet("Fig7_GenMix")
    _title(ws, "Fig 7 — Average intraday generation mix + net flow + price",
           "MW by UTC hour (pumped consumption shown negative) + price EUR/MWh · single-panel "
           "blocks ordered YEAR then country (future years append at the bottom) · 28-row stride")
    df = _load("intraday_genmix")
    series = [f"gen_{t}" for t in TECHS] + ["pumped_consumption", "flow_net", "price"]
    STRIDE = 28  # header + 24 hours + gap
    r = 4
    for y in DATA_YEARS:                      # YEAR outer -> future years append at bottom
        for country in COUNTRIES:
            ws.cell(r, 1, f"{cfg.COUNTRIES[country]['name']} {y}").font = Font(bold=True, color=NAVY)
            if y == CUR: ws.cell(r, 1).fill = YTD
            hdr = r + 1
            ws.cell(hdr, 1, "Hour UTC").font = H2; ws.cell(hdr, 1).fill = FILL_NAVY
            for k, s in enumerate(series):
                lab = s.replace("gen_", "").replace("_", " ")
                cell = ws.cell(hdr, 2 + k, lab); cell.font = H2; cell.fill = FILL_NAVY
                cell.alignment = Alignment(text_rotation=90, horizontal="center")
            sub = df[(df.country == country) & (df.year == y)].set_index("hour_utc")
            for h in range(24):
                rr = hdr + 1 + h
                ws.cell(rr, 1, f"{h:02d}h")
                if h in sub.index:
                    for k, s in enumerate(series):
                        val = sub.loc[h, s] if s in sub.columns else None
                        ws.cell(rr, 2 + k, round(float(val), 1) if pd.notna(val) else None)
            r += STRIDE
    ws.column_dimensions["A"].width = 12

def sheet_capture_monthly(wb):
    ws = wb.create_sheet("CaptureMonthly")
    _title(ws, "Capture price by technology — monthly",
           "absolute capture price EUR/MWh · rows = month · UTC")
    df = _load("capture_monthly")
    months = []
    for y in YEARS:
        for mo in range(1, 13):
            months.append((y, mo))
    # header: country (merged) then tech
    r0 = 4
    c = 2
    colmap = {}
    for country in COUNTRIES:
        ws.cell(r0, c, cfg.COUNTRIES[country]["name"]).font = H2; ws.cell(r0, c).fill = FILL_NAVY
        ws.merge_cells(start_row=r0, start_column=c, end_row=r0, end_column=c + len(TECHS) - 1)
        for t in TECHS:
            cell = ws.cell(r0 + 1, c, t); cell.font = Font(size=7, color="FFFFFF"); cell.fill = FILL_TEAL
            cell.alignment = Alignment(text_rotation=90)
            colmap[(country, t)] = c; c += 1
    ws.cell(r0 + 1, 1, "Month").font = BOLD
    for i, (y, mo) in enumerate(months):
        rr = r0 + 2 + i
        cell = ws.cell(rr, 1, f"{y}-{mo:02d}")
        if y == CUR: cell.fill = YTD
        for (country, t), col in colmap.items():
            v = df[(df.country == country) & (df.year == y) & (df.month == mo) & (df.tech == t)]["capture_price"]
            ws.cell(rr, col, round(float(v.iloc[0]), 2) if len(v) and pd.notna(v.iloc[0]) else None)
    ws.column_dimensions["A"].width = 10

def main():
    updated = pd.Timestamp.now(tz="UTC").strftime("%Y-%m-%d %H:%M UTC")
    wb = Workbook()
    sheet_readme(wb, updated)
    sheet_price_sd(wb)
    sheet_intraday_price(wb)
    sheet_neg_hours(wb)
    sheet_cum_neg(wb)
    sheet_duration(wb)
    sheet_capture(wb)
    sheet_daily_minmax(wb)
    sheet_genmix(wb)
    sheet_capacity(wb)
    sheet_capture_monthly(wb)
    wb.save(OUT)
    print(f"workbook -> {OUT}  ({len(wb.sheetnames)} tabs: {wb.sheetnames})", flush=True)

if __name__ == "__main__":
    main()
