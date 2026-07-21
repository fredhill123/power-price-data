"""
build_frozen_excel.py — a FROZEN snapshot of the live workbook: identical 16-sheet
structure and styled charts, but the data is hardcoded into the cells and ALL Power
Query is stripped out. Charts read from the static cells (Excel recalcs on open), so
it's a drop-in for a given month with zero live pulls — the linked deck can even be
re-pointed at this file instead of the live one.

Method (byte-surgery, never openpyxl-save — that would mangle the Redburn chart XML):
  1. base = the live workbook (styled charts + structure).
  2. rewrite each data sheet's <sheetData> from the fresh chart CSVs (+ fill G1/G2).
  3. strip PQ: tableParts, table/queryTable parts + rels, connections.xml, customXml,
     ExternalData_1 defined names, and their content-type overrides.

Usage: python build_frozen_excel.py [base.xlsx] [out.xlsx]
"""
from __future__ import annotations
import os, re, sys, zipfile, warnings; warnings.filterwarnings("ignore")
import pandas as pd
from lxml import etree
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "outputs", "HourlyPowerData.xlsx")
OUT  = sys.argv[2] if len(sys.argv) > 2 else os.path.join(ROOT, "outputs", "HourlyPowerData_frozen.xlsx")
CH   = os.path.join(ROOT, "outputs", "csv", "charts")
SS   = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RNS  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

# sheet name -> chart CSV that fills it
SHEET_CSV = {
    "Fig1_PriceSD": "fig1_price_sd", "Fig2_Intraday": "fig2_intraday_indexed",
    "Fig3_NegHours": "fig3_neg_hours_annual", "Fig3_CumNeg": "fig3_cum_near_neg",
    "Fig4_Duration": "fig4_duration_curve", "Fig5_Capture": "fig5_capture_pct",
    "Fig6_MinMax": "fig6_daily_minmax", "Fig7_GenMix": "fig7_gen_mix",
    "Fig9_Capacity": "fig9_capacity", "Fig2_Intraday_avg": "fig2_intraday_avg",
    "Fig5_Capture_abs": "fig5_capture_abs", "CaptureMonthly": "capture_monthly",
    "G1_SolarPeak": "g1_solar_peakhour", "G2_MonthDuck": "g2_price_by_month",
    "A_MonthPrice": "figA_monthly_price", "B_Penetration": "figB_penetration",
    "C_CaptureErosion": "figC_capture_erosion", "D_NetloadDuck": "figD_netload_duck",
}

def esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

def sheetdata_xml(df):
    """Build <sheetData> from a dataframe: row1 = string headers, rows 2+ = values."""
    out = ["<sheetData>"]
    # header
    out.append('<row r="1">')
    for j, col in enumerate(df.columns):
        ref = f"{get_column_letter(j+1)}1"
        out.append(f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{esc(col)}</t></is></c>')
    out.append("</row>")
    # data
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        cells = []
        for j, v in enumerate(row.values):
            if v is None or pd.isna(v):
                continue
            ref = f"{get_column_letter(j+1)}{i}"
            if isinstance(v, str):
                cells.append(f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{esc(v)}</t></is></c>')
            else:                                   # numeric (incl. numpy scalars)
                fv = float(v)
                txt = str(int(fv)) if fv.is_integer() else repr(fv)
                cells.append(f'<c r="{ref}"><v>{txt}</v></c>')
        out.append(f'<row r="{i}">' + "".join(cells) + "</row>")
    out.append("</sheetData>")
    return "".join(out)

def main():
    zin = zipfile.ZipFile(BASE)
    parts = {i.filename: zin.read(i.filename) for i in zin.infolist()}
    order = [i.filename for i in zin.infolist()]
    zin.close()

    # --- map sheet name -> worksheet file ---
    wbx = etree.fromstring(parts["xl/workbook.xml"]); ns = {"m": SS, "r": RNS}
    rid_target = {r.get("Id"): r.get("Target")
                  for r in etree.fromstring(parts["xl/_rels/workbook.xml.rels"])}
    name_file = {}
    for s in wbx.findall(".//m:sheets/m:sheet", ns):
        rid = s.get(f"{{{RNS}}}id")
        name_file[s.get("name")] = "xl/" + rid_target[rid]

    # --- 1. rewrite sheetData for each data sheet from fresh CSVs ---
    for name, csv in SHEET_CSV.items():
        f = name_file[name]
        df = pd.read_csv(os.path.join(CH, f"{csv}.csv"))
        xml = parts[f].decode()
        new_sd = sheetdata_xml(df)
        if "<sheetData/>" in xml:
            xml = xml.replace("<sheetData/>", new_sd)
        else:
            xml = re.sub(r"<sheetData>.*?</sheetData>", new_sd, xml, flags=re.DOTALL)
        xml = re.sub(r"<tableParts.*?</tableParts>", "", xml, flags=re.DOTALL)   # drop table link
        nrows, ncols = df.shape
        xml = re.sub(r'<dimension ref="[^"]*"/>',
                     f'<dimension ref="A1:{get_column_letter(ncols)}{nrows+1}"/>', xml)
        parts[f] = xml.encode()
        # sheet rels: drop the table relationship
        rf = f.replace("worksheets/", "worksheets/_rels/") + ".rels"
        if rf in parts:
            rr = re.sub(r'<Relationship[^>]*Type="[^"]*/table"[^>]*/>', "", parts[rf].decode())
            parts[rf] = rr.encode()

    # --- 2. strip PQ parts ---
    drop = [n for n in list(parts) if re.match(r"xl/tables/", n) or re.match(r"xl/queryTables/", n)
            or n == "xl/connections.xml" or n.startswith("customXml/")]
    for n in drop:
        parts.pop(n, None)
    order = [n for n in order if n not in drop]

    # workbook.xml: drop ExternalData_1 defined names
    wb = parts["xl/workbook.xml"].decode()
    wb = re.sub(r"<definedNames>.*?</definedNames>", "", wb, flags=re.DOTALL)
    parts["xl/workbook.xml"] = wb.encode()
    # workbook rels: drop connections + customXml
    wr = parts["xl/_rels/workbook.xml.rels"].decode()
    wr = re.sub(r'<Relationship[^>]*Type="[^"]*/(connections|customXml)"[^>]*/>', "", wr)
    parts["xl/_rels/workbook.xml.rels"] = wr.encode()
    # content types: drop overrides for the stripped parts
    ct = parts["[Content_Types].xml"].decode()
    ct = re.sub(r'<Override PartName="/xl/(tables/table\d+|queryTables/queryTable\d+|connections)\.xml"[^>]*/>', "", ct)
    ct = re.sub(r'<Override PartName="/customXml/[^"]*"[^>]*/>', "", ct)
    parts["[Content_Types].xml"] = ct.encode()

    # --- 3. write ---
    seen = set(); final = []
    for n in order:
        if n in parts and n not in seen:
            seen.add(n); final.append(n)
    zout = zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED)
    for n in final:
        zout.writestr(n, parts[n])
    zout.close()
    print(f"wrote frozen workbook -> {OUT}  ({len(final)} parts, PQ stripped, {len(SHEET_CSV)} sheets hardcoded)")

if __name__ == "__main__":
    main()
