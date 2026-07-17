"""
build_linked.py — build the deliverable workbook: pre-built charts, NO seed data.

Each figure sheet ships EMPTY except for a floating chart whose series point at
the range where Fred's Power Query will load (headers row 1, data below, out to
2035). Loading a query into empty cells fills them IN PLACE with no shift, so the
chart reads the live query — unlike a seeded target, which Power Query shoves
aside (displacing the chart's references onto stale copies).

Rule: Claude builds charts (openpyxl), Fred adds queries in Excel (preserves the
charts). Never the reverse — openpyxl would strip the Power Query.

SEED=True embeds the data for a local render check only; delivery uses SEED=False.

Reads : outputs/csv/charts/*.csv   (for column layout / dimensions)
Writes: outputs/PowerPriceData_Linked.xlsx
"""
from __future__ import annotations
import os, sys, warnings; warnings.filterwarnings("ignore")
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import config as cfg

CH = os.path.join(cfg.OUTPUT_DIR, "csv", "charts")
OUT = os.path.join(cfg.OUTPUT_DIR, "PowerPriceData_Linked.xlsx")
SEED = "--seed" in sys.argv

NAVY = "1F3864"; TEAL = "2E7D8A"
YEAR_RAMP = ["1F3864", "27496D", "2E5C7A", "357084", "3C8391",
             "5FA1AD", "8FBEC6", "B9D3D8"]
COUNTRY_COLORS = ["1F3864", "8A1E41", "CC9F53", "2E7D8A", "3D664A"]
BASE = "https://raw.githubusercontent.com/fredhill123/power-price-data/main/published/charts/"
SETUP = []  # (sheet, csv, load_target) rows for READ_ME

def load(name): return pd.read_csv(os.path.join(CH, f"{name}.csv"))

def prep(wb, sheet, df, csvname):
    """Create the sheet; write data only if SEED. Returns (ws, nrows, ncols)."""
    ws = wb.create_sheet(sheet)
    nrows, ncols = df.shape           # nrows = data rows (excl header)
    if SEED:
        ws.append(list(df.columns))
        for _, r in df.iterrows():
            ws.append([None if pd.isna(v) else v for v in r.values])
    SETUP.append((sheet, csvname, f"='{sheet}'!$A$1"))
    return ws, nrows, ncols

def anchor(nrows):
    return f"A{nrows + 3}"          # chart sits below the (empty) data region

def col_of(df, name): return list(df.columns).index(name) + 1

def style_line(ch, colors, w=20000):
    for i, s in enumerate(ch.series):
        s.graphicalProperties = GraphicalProperties()
        s.graphicalProperties.line = LineProperties(solidFill=colors[i % len(colors)], w=w)
        s.smooth = False

def style_bar(ch, colors):
    for i, s in enumerate(ch.series):
        s.graphicalProperties = GraphicalProperties(solidFill=colors[i % len(colors)])

# ---- figures ---------------------------------------------------------------
def fig1(wb):
    df = load("fig1_price_sd"); ws, n, _ = prep(wb, "Fig1_PriceSD", df, "fig1_price_sd")
    ch = LineChart(); ch.title = "Fig 1 - Price volatility (SD) by year"; ch.y_axis.title = "EUR/MWh"
    ch.height = 9; ch.width = 17
    ch.add_data(Reference(ws, min_col=2, max_col=6, min_row=1, max_row=n + 1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    style_line(ch, COUNTRY_COLORS); ws.add_chart(ch, anchor(n))

def year_series(wb, sheet, csv, country, title, ytitle, kind="line"):
    df = load(csv); ws, n, _ = prep(wb, sheet, df, csv)
    cols = [f"{country}_{y}" for y in cfg.YEARS if f"{country}_{y}" in df.columns]
    ch = (LineChart() if kind == "line" else BarChart())
    if kind == "bar": ch.type = "col"; ch.grouping = "clustered"
    ch.title = title; ch.y_axis.title = ytitle; ch.height = 9; ch.width = 18
    for c in cols:
        ch.add_data(Reference(ws, min_col=col_of(df, c), min_row=1, max_row=n + 1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    (style_line if kind == "line" else style_bar)(ch, YEAR_RAMP)
    ws.add_chart(ch, anchor(n))

def fig3(wb):
    df = load("fig3_neg_hours_annual"); ws, n, _ = prep(wb, "Fig3_NegHours", df, "fig3_neg_hours_annual")
    ch = BarChart(); ch.type = "col"; ch.grouping = "clustered"
    ch.title = "Fig 3 - Negative price hours by year"; ch.y_axis.title = "# hours"; ch.height = 9; ch.width = 17
    for c in [c for c in df.columns if c.endswith("_neg")]:
        ch.add_data(Reference(ws, min_col=col_of(df, c), min_row=1, max_row=n + 1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    style_bar(ch, COUNTRY_COLORS); ws.add_chart(ch, anchor(n))

def fig6(wb):
    df = load("fig6_daily_minmax"); ws, n, _ = prep(wb, "Fig6_MinMax", df, "fig6_daily_minmax")
    ch = ScatterChart(); ch.title = "Fig 6 - Daily min vs max price (Germany 2024)"
    ch.x_axis.title = "max EUR/MWh"; ch.y_axis.title = "min EUR/MWh"; ch.height = 11; ch.width = 13
    xr = Reference(ws, min_col=col_of(df, "DE_2024_max"), min_row=2, max_row=n + 1)
    yr = Reference(ws, min_col=col_of(df, "DE_2024_min"), min_row=2, max_row=n + 1)
    s = Series(yr, xr, title="DE 2024"); s.marker.symbol = "circle"; s.marker.size = 4
    s.graphicalProperties = GraphicalProperties(); s.graphicalProperties.line = LineProperties(noFill=True)
    s.marker.graphicalProperties = GraphicalProperties(solidFill=TEAL)
    ch.series.append(s); ws.add_chart(ch, anchor(n))

def fig7(wb):
    df = load("fig7_gen_mix"); ws, n, _ = prep(wb, "Fig7_GenMix", df, "fig7_gen_mix")
    barcols = [f"PT_2024_{t}" for t in cfg.TECH_ORDER if f"PT_2024_{t}" in df.columns]
    bar = BarChart(); bar.type = "col"; bar.grouping = "stacked"; bar.overlap = 100
    bar.title = "Fig 7 - Intraday generation mix + price (Portugal 2024)"; bar.y_axis.title = "MW"
    bar.height = 11; bar.width = 20
    for c in barcols:
        bar.add_data(Reference(ws, min_col=col_of(df, c), min_row=1, max_row=n + 1), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    if "PT_2024_price" in df.columns:
        line = LineChart()
        line.add_data(Reference(ws, min_col=col_of(df, "PT_2024_price"), min_row=1, max_row=n + 1), titles_from_data=True)
        line.y_axis.axId = 200; line.y_axis.crosses = "max"; style_line(line, ["000000"], w=26000)
        bar += line
    ws.add_chart(bar, anchor(n))

def data_only(wb, sheet, csv):
    df = load(csv); prep(wb, sheet, df, csv)

def readme(wb):
    ws = wb.create_sheet("READ_ME_FIRST", 0)
    hdr = [
        ("Power Price Data - linked workbook", 14, NAVY, True),
        ("", 10, "000000", False),
        ("Each sheet has a chart already built. On each sheet, add ONE Power Query:", 10, "000000", True),
        ("  Data > Get Data > From Web > paste the URL below > Load To... > Existing worksheet > that sheet's $A$1", 10, "000000", False),
        ("IMPORTANT: the data cells are intentionally EMPTY - load the query into them (do not pre-type anything there).", 10, "C00000", True),
        ("Then Data > Queries & Connections > each query > Properties > tick 'Refresh data when opening the file'.", 10, "C00000", True),
        ("Charts are pre-wired and pre-allocated to 2035, so future years fill in automatically on refresh.", 10, "808080", False),
        ("", 10, "000000", False),
    ]
    r = 1
    for t, sz, col, b in hdr:
        ws.cell(r, 1, t).font = Font(size=sz, color=col, bold=b); r += 1
    ws.cell(r, 1, "Sheet").font = Font(bold=True); ws.cell(r, 2, "From Web URL").font = Font(bold=True)
    ws.cell(r, 3, "Load to").font = Font(bold=True); r += 1
    for sheet, csv, tgt in SETUP:
        ws.cell(r, 1, sheet); ws.cell(r, 2, BASE + csv + ".csv"); ws.cell(r, 3, tgt); r += 1
    ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 78; ws.column_dimensions["C"].width = 22

def main():
    wb = Workbook(); wb.remove(wb.active)
    fig1(wb)
    year_series(wb, "Fig2_Intraday", "fig2_intraday_indexed", "DE", "Fig 2 - Indexed intraday price (Germany)", "indexed (1=base)", "line")
    fig3(wb)
    year_series(wb, "Fig3_CumNeg", "fig3_cum_near_neg", "DE", "Fig 3 - Cumulative near-negative hours (Germany)", "# hours (cum)", "line")
    year_series(wb, "Fig4_Duration", "fig4_duration_curve", "PT", "Fig 4 - Price duration curves (Portugal)", "EUR/MWh", "line")
    year_series(wb, "Fig5_Capture", "fig5_capture_pct", "DE", "Fig 5 - Capture price vs base by tech (Germany)", "% vs base", "bar")
    fig6(wb); fig7(wb)
    year_series(wb, "Fig9_Capacity", "fig9_capacity", "DE", "Fig 9 - Installed capacity by tech (Germany)", "MW", "bar")
    data_only(wb, "Fig2_Intraday_avg", "fig2_intraday_avg")
    data_only(wb, "Fig5_Capture_abs", "fig5_capture_abs")
    data_only(wb, "CaptureMonthly", "capture_monthly")
    readme(wb)
    wb.save(OUT)
    print(f"saved {OUT}  SEED={SEED}  ({len(wb.sheetnames)} sheets)", flush=True)

if __name__ == "__main__":
    main()
